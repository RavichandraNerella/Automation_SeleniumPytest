""" This module is used for developing/ accessing data reader utility. """

import os

import openpyxl

from Utils.Logs.logger import logger


class ExcelReader:

    def __init__(self):
        self.cur_path = os.path.abspath(os.getcwd())

    def load_test_data(self, excel_name, sheet_name):
        """
        This method is used for loading excel file data
        :return: it returns excel records
        """
        dict_list = []

        ui_file_path = self.cur_path + "/TestData/{}.xlsx".format(excel_name)

        try:
            if ui_file_path is not None:
                wb = openpyxl.load_workbook(ui_file_path)
                sheet = wb[sheet_name]

                col_names = list()
                j = 0
                for COL in sheet.iter_cols(1, sheet.max_column):
                    col_names.append(COL[j].value)

                for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    row_data = {}
                    count = 0
                    for cell in row_cells:
                        row_data[col_names[count]] = cell.value
                        count = count + 1
                    dict_list.append(row_data)
        except Exception as ex:
            logger.exception(ex)
            raise Exception(ex)

        return dict_list

    def get_column_data(self, excel_name, column_name, row_number):
        """
        This method is used for returning specific column data
        :param excel_name: it takes the name of the Excel workbook for which data has to be returned
        :param column_name: it takes the name of the column for which value has to be returned
        :param row_number: it takes the row number of the sheet for which column value has to be returned
        :return: value
        """
        value = None
        excel_records = self.load_test_data(excel_name)

        # noinspection PyBroadException

        try:
            value = excel_records[row_number][column_name]

        except Exception as ex:
            logger.exception(ex)
            raise Exception(ex)

        return value

    def get_data(self, excel_name, sheet_name):
        """
        This method is used for returning all the records of the sheet
        :param excel_name: it takes the name of the Excel workbook for which data has to be returned
        :param sheet_name: it takes the name of the Excel sheet name for which data has to be returned
        :return: records
        """
        try:
            excel_records = self.load_test_data(excel_name, sheet_name)

        except Exception as ex:
            logger.exception(ex)
            raise Exception(ex)

        return excel_records
